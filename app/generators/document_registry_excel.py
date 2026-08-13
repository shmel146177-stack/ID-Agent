from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.services.document_completeness import (
    document_completeness,
)
from app.services.document_registry import (
    document_registry,
)


class DocumentRegistryExcel:
    """
    Формирование Excel-отчёта по проекту.

    Листы:
    1. Реестр документов
    2. Комплектность

    Для проектной документации комплектность
    определяется по ведомости рабочих чертежей
    и фактически найденным страницам PDF.
    """

    def __init__(self):

        self.header_fill = PatternFill(
            "solid",
            fgColor="1F4E78",
        )

        self.subheader_fill = PatternFill(
            "solid",
            fgColor="D9EAF7",
        )

        self.success_fill = PatternFill(
            "solid",
            fgColor="E2F0D9",
        )

        self.warning_fill = PatternFill(
            "solid",
            fgColor="FFF2CC",
        )

        self.error_fill = PatternFill(
            "solid",
            fgColor="F4CCCC",
        )

        self.white_font = Font(
            color="FFFFFF",
            bold=True,
        )

        self.bold_font = Font(
            bold=True,
        )

        self.thin_side = Side(
            style="thin",
            color="B7B7B7",
        )

        self.border = Border(
            left=self.thin_side,
            right=self.thin_side,
            top=self.thin_side,
            bottom=self.thin_side,
        )

    def _project_path(
        self,
        project_name: str,
    ) -> Path:

        return Path("projects") / project_name

    def _output_path(
        self,
        project_name: str,
    ) -> Path:

        return (
            self._project_path(project_name)
            / "output"
            / ("Реестр_документов_" f"{project_name}.xlsx")
        )

    def _style_header(
        self,
        worksheet,
        row_number: int,
        columns_count: int,
    ):

        for column in range(
            1,
            columns_count + 1,
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column,
            )

            cell.fill = self.header_fill

            cell.font = self.white_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = self.border

    def _style_table(
        self,
        worksheet,
        start_row: int,
        end_row: int,
        columns_count: int,
    ):

        for row in worksheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=1,
            max_col=columns_count,
        ):

            for cell in row:

                cell.border = self.border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    def _set_column_widths(
        self,
        worksheet,
        widths: dict,
    ):

        for column, width in widths.items():

            worksheet.column_dimensions[column].width = width

    def _create_registry_sheet(
        self,
        workbook: Workbook,
        project_name: str,
        registry: dict,
    ):

        worksheet = workbook.active

        worksheet.title = "Реестр документов"

        # ---------------------------------------------------------
        # ЗАГОЛОВОК
        # ---------------------------------------------------------

        worksheet.merge_cells("A1:I1")

        worksheet["A1"] = "РЕЕСТР ДОКУМЕНТОВ"

        worksheet["A1"].font = Font(
            bold=True,
            size=14,
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.merge_cells("A2:I2")

        worksheet["A2"] = f"Проект: {project_name}"

        worksheet["A2"].font = self.bold_font

        # ---------------------------------------------------------
        # ТАБЛИЦА
        # ---------------------------------------------------------

        headers = [
            "№",
            "Файл",
            "Тип документа",
            "Статус",
            "Расширение",
            "Номер чертежа",
            "Дата",
            "Изготовитель",
            "Оборудование",
        ]

        header_row = 4

        for column, value in enumerate(
            headers,
            start=1,
        ):

            worksheet.cell(
                row=header_row,
                column=column,
                value=value,
            )

        self._style_header(
            worksheet,
            header_row,
            len(headers),
        )

        documents = registry.get(
            "documents",
            [],
        )

        row_number = header_row + 1

        for document in documents:

            values = [
                document.get("number"),
                document.get("filename"),
                document.get("classification"),
                document.get("status"),
                document.get("extension"),
                document.get("drawing_number"),
                document.get("date"),
                document.get("manufacturer"),
                document.get("equipment"),
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):

                worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

            row_number += 1

        if documents:

            self._style_table(
                worksheet,
                header_row + 1,
                row_number - 1,
                len(headers),
            )

        worksheet.freeze_panes = "A5"

        worksheet.auto_filter.ref = f"A4:I{max(row_number - 1, 4)}"

        self._set_column_widths(
            worksheet,
            {
                "A": 7,
                "B": 38,
                "C": 22,
                "D": 16,
                "E": 12,
                "F": 25,
                "G": 15,
                "H": 28,
                "I": 35,
            },
        )

    def _create_completeness_sheet(
        self,
        workbook: Workbook,
        project_name: str,
        completeness: dict,
    ):

        worksheet = workbook.create_sheet("Комплектность")

        # ---------------------------------------------------------
        # ЗАГОЛОВОК
        # ---------------------------------------------------------

        worksheet.merge_cells("A1:H1")

        worksheet["A1"] = "ПРОВЕРКА КОМПЛЕКТНОСТИ"

        worksheet["A1"].font = Font(
            bold=True,
            size=14,
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.merge_cells("A2:H2")

        worksheet["A2"] = f"Проект: {project_name}"

        worksheet["A2"].font = self.bold_font

        # ---------------------------------------------------------
        # СВОДКА
        # ---------------------------------------------------------

        summary = [
            (
                "Метод проверки",
                completeness.get(
                    "check_method",
                    completeness.get(
                        "profile_name",
                        "",
                    ),
                ),
            ),
            (
                "Статус",
                completeness.get(
                    "status",
                    "",
                ),
            ),
            (
                "Требуется листов",
                completeness.get(
                    "required_count",
                    0,
                ),
            ),
            (
                "Найдено листов",
                completeness.get(
                    "found_count",
                    0,
                ),
            ),
            (
                "Отсутствует листов",
                completeness.get(
                    "missing_count",
                    0,
                ),
            ),
            (
                "Комплектность",
                (f"{completeness.get('completeness_percent', 0)}%"),
            ),
        ]

        start_summary_row = 4

        for index, (
            label,
            value,
        ) in enumerate(
            summary,
            start=start_summary_row,
        ):

            worksheet.cell(
                row=index,
                column=1,
                value=label,
            )

            worksheet.cell(
                row=index,
                column=2,
                value=value,
            )

            worksheet.cell(
                row=index,
                column=1,
            ).font = self.bold_font

            worksheet.cell(
                row=index,
                column=1,
            ).fill = self.subheader_fill

            for column in (
                1,
                2,
            ):

                worksheet.cell(
                    row=index,
                    column=column,
                ).border = self.border

                worksheet.cell(
                    row=index,
                    column=column,
                ).alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        status_cell = worksheet.cell(
            row=start_summary_row + 1,
            column=2,
        )

        if (
            completeness.get(
                "missing_count",
                0,
            )
            == 0
        ):

            status_cell.fill = self.success_fill

        else:

            status_cell.fill = self.error_fill

        # ---------------------------------------------------------
        # ТАБЛИЦА ЛИСТОВ
        # ---------------------------------------------------------

        table_row = start_summary_row + len(summary) + 2

        headers = [
            "№ листа",
            "Наименование",
            "Статус",
            "PDF-страница",
            "Файл",
            "Тип страницы",
            "Балл",
            "Уверенность",
        ]

        for column, value in enumerate(
            headers,
            start=1,
        ):

            worksheet.cell(
                row=table_row,
                column=column,
                value=value,
            )

        self._style_header(
            worksheet,
            table_row,
            len(headers),
        )

        documents = completeness.get(
            "documents",
            [],
        )

        row_number = table_row + 1

        for document in documents:

            values = [
                document.get("sheet_number"),
                (document.get("title") or document.get("document_type")),
                document.get("status"),
                document.get("matched_page"),
                document.get("matched_filename"),
                document.get("matched_page_type"),
                document.get("score"),
                document.get("confidence"),
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):

                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

                cell.border = self.border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            status = document.get("status")

            if status == "Есть":

                worksheet.cell(
                    row=row_number,
                    column=3,
                ).fill = self.success_fill

            else:

                for column in range(
                    1,
                    len(headers) + 1,
                ):

                    worksheet.cell(
                        row=row_number,
                        column=column,
                    ).fill = self.error_fill

            row_number += 1

        # ---------------------------------------------------------
        # ОТСУТСТВУЮЩИЕ ЛИСТЫ
        # ---------------------------------------------------------

        missing_sheets = completeness.get(
            "missing_sheets",
            [],
        )

        missing_start = row_number + 2

        worksheet.merge_cells(
            start_row=missing_start,
            start_column=1,
            end_row=missing_start,
            end_column=8,
        )

        worksheet.cell(
            row=missing_start,
            column=1,
            value="Отсутствующие листы",
        )

        worksheet.cell(
            row=missing_start,
            column=1,
        ).font = self.bold_font

        worksheet.cell(
            row=missing_start,
            column=1,
        ).fill = self.warning_fill

        if missing_sheets:

            current_row = missing_start + 1

            for missing in missing_sheets:

                worksheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=8,
                )

                worksheet.cell(
                    row=current_row,
                    column=1,
                    value=(
                        f"Лист №{missing.get('sheet_number')}: "
                        f"{missing.get('title')}"
                    ),
                )

                worksheet.cell(
                    row=current_row,
                    column=1,
                ).fill = self.error_fill

                worksheet.cell(
                    row=current_row,
                    column=1,
                ).border = self.border

                current_row += 1

        else:

            worksheet.merge_cells(
                start_row=missing_start + 1,
                start_column=1,
                end_row=missing_start + 1,
                end_column=8,
            )

            worksheet.cell(
                row=missing_start + 1,
                column=1,
                value="Отсутствующие листы не обнаружены",
            )

            worksheet.cell(
                row=missing_start + 1,
                column=1,
            ).fill = self.success_fill

        worksheet.freeze_panes = f"A{table_row + 1}"

        self._set_column_widths(
            worksheet,
            {
                "A": 12,
                "B": 42,
                "C": 16,
                "D": 15,
                "E": 34,
                "F": 24,
                "G": 10,
                "H": 16,
            },
        )

        # Центрируем служебные столбцы.
        for row in worksheet.iter_rows(
            min_row=table_row + 1,
            max_row=max(
                row_number - 1,
                table_row + 1,
            ),
        ):

            for column_index in (
                1,
                3,
                4,
                7,
                8,
            ):

                row[column_index - 1].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    def create(
        self,
        project_name: str,
    ) -> str:

        registry = document_registry.build(project_name)

        completeness = document_completeness.check(project_name)

        workbook = Workbook()

        self._create_registry_sheet(
            workbook,
            project_name,
            registry,
        )

        self._create_completeness_sheet(
            workbook,
            project_name,
            completeness,
        )

        output_path = self._output_path(project_name)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(output_path)

        return str(output_path)


document_registry_excel = DocumentRegistryExcel()
